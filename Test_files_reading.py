from zipfile import ZipFile
import csv

from PyPDF2 import PdfReader
from openpyxl import load_workbook

ARCHIVE_PATH = "resources/files_arc.zip"

def test_archive_has_required_types():
    with ZipFile(ARCHIVE_PATH, "r") as zf:
        names = zf.namelist()
    assert any(n.lower().endswith(".csv") for n in names), "В архиве нет CSV"
    assert any(n.lower().endswith(".pdf") for n in names), "В архиве нет PDF"
    assert any(n.lower().endswith(".xlsx") for n in names), "В архиве нет XLSX"

def test_read_csv_from_zip():
    with ZipFile(ARCHIVE_PATH, "r") as zf:
        # читаем текстом с utf-8
        with zf.open("example1.csv") as csv_file:
            content = csv_file.read().decode('utf-8')
            # проверка
            assert len(content) >= 1, "CSV без строк данных"

def test_read_pdf_from_zip():
    with ZipFile(ARCHIVE_PATH, "r") as zf:
        with zf.open("QA_JS.pdf") as pdf_file:
            pdf = PdfReader(pdf_file)
            # проверка
            assert len(pdf.pages) >= 1, "PDF пустой (нет страниц)"


def test_read_xlsx_from_zip():
    with ZipFile(ARCHIVE_PATH, "r") as zf:
        with zf.open("xls_test.xlsx") as xls_file:
             wb = load_workbook(xls_file)
             ws = wb[wb.sheetnames[0]]
        # проверка
        assert ws.max_row >= 1 and ws.max_column >= 1, "XLSX выглядит пустым"
